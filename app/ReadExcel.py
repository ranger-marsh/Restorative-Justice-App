
'''
Handles the opening, creation, reading and writing of excel files. As it
copies data from an excel file it does some basic sanitation specific to
this application.
'''

from openpyxl import load_workbook


class ReadExcel:

    def __init__(self, path=''):
        self.path = path
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active
        self.max_row = self.ws.max_row
        self.max_col = self.ws.max_column

    def validate(self):
        ws_heders = self.copy_row(1)
        expected = ['case number',
                    'case occurred from date',
                    'case occurred incident type',
                    'case ori',
                    'case subject age',
                    'case subject custody status',
                    'case subject global subject',
                    'case subject global subject address',
                    'case subject global subject address apartment',
                    'case subject global subject address city',
                    'case subject global subject address state',
                    'case subject global subject date of birth',
                    'case subject global subject primary phone number',
                    'case subject global subject race',
                    'case subject global subject sex',
                    'case subject type',
                    'reporting district'
                    ]
        if ws_heders == expected:
            return True
        return False

    def copy_row(self, row_to_copy):
        # copy a given row into a list() each cell value is a list value
        row_list = list()
        for col in range(1, (self.max_col + 1)):
            row_list.append(self.ws.cell(row=row_to_copy, column=col).value)
        row_list = self.clean_list(row_list)
        return row_list

    def copy_col(self, col_to_copy, start_row):
        # copy a given column into a list() each cell value is a list value
        col_list = list()
        for row in range(start_row, (self.max_row + 1)):
            col_list.append(self.ws.cell(row=row, column=col_to_copy).value)
            col_list = self.clean_list(col_list)
        return col_list

    def clean_list(self, dirty_list):
        # sanitize a list so that all items are lowercase strings
        clean_list = [self.clean_str(item) for item in dirty_list]
        return clean_list

    def clean_str(self, dirty_str):
        # sanitize a value so that it is a lowercase string
        if dirty_str is None:
            clean_string = ''
        else:
            dirty_str = str(dirty_str)
            clean_string = dirty_str.strip().lower()
        return clean_string
